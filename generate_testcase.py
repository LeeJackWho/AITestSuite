import time
import os
import re
import json
import pandas as pd
import requests
import argparse
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from dotenv import load_dotenv

# 加载环境变量
load_dotenv()

# 获取API密钥和配置
AI_BASE_URL = os.getenv("AI_BASE_URL")
AI_API_ENDPOINT = os.getenv("AI_API_ENDPOINT")
AI_API_KEY = os.getenv("AI_API_KEY")
MODEL_NAME = os.getenv("MODEL_NAME", "default")

# 确定API端点，优先使用AI_BASE_URL
API_ENDPOINT = AI_BASE_URL or AI_API_ENDPOINT

# 获取特定模型的API密钥（可选）
QIANWEN_API_KEY = os.getenv("QIANWEN_API_KEY")
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

# 通用模型配置
MODEL_CONFIGS = {
    "default": {
        "api_key_env": "AI_API_KEY",
        "endpoint": API_ENDPOINT,
        "headers": lambda key: {
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json; charset=utf-8"
        },
        "payload": lambda messages, temperature: {
            "model": MODEL_NAME,
            "messages": messages,
            "temperature": temperature
        },
        "response_parser": lambda json_data: json_data["choices"][0]["message"]["content"] if "choices" in json_data and len(json_data["choices"]) > 0 else ""
    }
}

# 如果有特定模型的API密钥，添加对应的配置
if QIANWEN_API_KEY:
    MODEL_CONFIGS["qianwen"] = {
        "api_key_env": "QIANWEN_API_KEY",
        "endpoint": "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation",
        "headers": lambda key: {
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json; charset=utf-8"
        },
        "payload": lambda messages, temperature: {
            "model": "qwen-max",
            "input": {
                "messages": messages
            },
            "parameters": {
                "temperature": temperature,
                "result_format": "message"
            }
        },
        "response_parser": lambda json_data: json_data["output"]["message"]["content"] if "output" in json_data and "message" in json_data["output"] else ""
    }

# 添加Gemini模型配置
if GEMINI_API_KEY:
    MODEL_CONFIGS["mygemini"] = {
        "api_key_env": "GEMINI_API_KEY",
        "endpoint": os.getenv("GEMINI_BASE_URL"),  # 从环境变量获取GEMINI_BASE_URL
        "headers": lambda key: {
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json"
        },
        "payload": lambda messages, temperature: {
            "model": os.getenv("GEMINI_MODEL_NAME"),  # 从环境变量获取模型名称
            "messages": messages,
            "temperature": temperature  # 添加temperature参数
        },
        "response_parser": lambda json_data: json_data["choices"][0]["message"]["content"] if "choices" in json_data and len(json_data["choices"]) > 0 else ""
    }

# 添加OpenRouter模型配置
if OPENROUTER_API_KEY:
    MODEL_CONFIGS["myopenrouter"] = {
        "api_key_env": "OPENROUTER_API_KEY",
        "endpoint": os.getenv("OPENROUTER_BASE_URL"), 
        "headers": lambda key: {
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json"
        },
        "payload": lambda messages, temperature: {
            "model": os.getenv("OPENROUTER_MODEL_NAME"),  # 从环境变量获取模型名称
            "messages": messages,
            "temperature": temperature  # 添加temperature参数
        },
        "response_parser": lambda json_data: json_data["choices"][0]["message"]["content"] if "choices" in json_data and len(json_data["choices"]) > 0 else ""
    }

# 修改默认模型
DEFAULT_MODEL = "default"

# 默认配置
DEFAULT_CONFIG = {
    "需求分类": "测试需求",
    "迭代": "迭代27",
    "处理人": ""
}

# 优先级映射
PRIORITY_MAP = {
    "高": "High",
    "中": "Middle",
    "低": "Low",
    "可选": "Nice To Have",
    "high": "High",
    "middle": "Middle",
    "low": "Low",
    "nice to have": "Nice To Have"
}

def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='AI测试用例生成器')
    parser.add_argument('--input', type=str, help='输入需求文件路径')
    parser.add_argument('--model', type=str, default=DEFAULT_MODEL, help='使用的AI模型')
    parser.add_argument('--output-dir', type=str, default='./测试用例', help='测试用例输出目录')
    parser.add_argument('--report-dir', type=str, default='./测试报告', help='测试报告输出目录')
    return parser.parse_args()

def read_excel_requirements(file_path):
    """读取Excel需求文档"""
    try:
        df = pd.read_excel(file_path)
        
        # 检查必要的列是否存在
        required_columns = ["标题", "详细描述"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            print(f"错误：Excel文件缺少必要的列: {', '.join(missing_columns)}")
            return None
        
        # 如果没有需求ID列，添加自动生成的ID
        if "需求ID" not in df.columns:
            df["需求ID"] = [f"REQ{i+1:03d}" for i in range(len(df))]
        
        # 如果没有优先级列，添加默认值
        if "优先级" not in df.columns:
            df["优先级"] = "中"
        
        # 如果没有父需求列，添加空值
        if "父需求" not in df.columns:
            df["父需求"] = ""
            
        # 如果没有需求分类列，添加默认值
        if "需求分类" not in df.columns:
            df["需求分类"] = DEFAULT_CONFIG["需求分类"]
            
        # 如果没有迭代列，添加默认值
        if "迭代" not in df.columns:
            df["迭代"] = DEFAULT_CONFIG["迭代"]
            
        # 如果没有处理人列，添加默认值
        if "处理人" not in df.columns:
            df["处理人"] = DEFAULT_CONFIG["处理人"]
        
        # 转换为字典列表
        requirements = df.to_dict('records')
        return requirements
    except Exception as e:
        print(f"读取Excel文件失败: {str(e)}")
        return None

def call_ai_model(model_name, messages, max_retries=3, temperature=0.3):
    """调用AI模型生成内容"""
    if model_name not in MODEL_CONFIGS:
        print(f"错误：不支持的模型 '{model_name}'，将使用默认模型 '{DEFAULT_MODEL}'")
        model_name = DEFAULT_MODEL
    
    model_config = MODEL_CONFIGS[model_name]
    api_key = os.getenv(model_config["api_key_env"])
    
    if not api_key:
        print(f"错误：未设置 {model_config['api_key_env']} 环境变量")
        return None
    
    headers = model_config["headers"](api_key)
    endpoint = model_config["endpoint"]
    
    if not endpoint:
        print(f"错误：未设置API端点，请在.env文件中配置AI_BASE_URL或AI_API_ENDPOINT")
        return None
    
    total_tokens = 0
    
    for attempt in range(max_retries):
        try:
            payload = model_config["payload"](messages, temperature)
            
            # 将payload转换为JSON字符串，确保正确处理中文
            json_data_str = json.dumps(payload, ensure_ascii=False)
            
            # 准备请求参数
            request_kwargs = {
                "headers": headers,
                "data": json_data_str.encode('utf-8'),
                "timeout": 60
            }
            
            # 如果有URL参数，添加到请求中
            if "url_params" in model_config:
                url_params = model_config["url_params"](api_key)
                endpoint = f"{endpoint}?{'&'.join([f'{k}={v}' for k, v in url_params.items()])}"
            
            # 发送请求
            print(f"正在调用API: {endpoint}")
            response = requests.post(endpoint, **request_kwargs)
            
            # 检查响应状态
            response.raise_for_status()
            
            # 解析响应
            json_data = response.json()
            
            # 使用模型特定的解析器提取内容
            content = model_config["response_parser"](json_data)
            
            if not content:
                print(f"警告：模型 {model_name} 返回的内容为空")
                
            if "usage" in json_data and "total_tokens" in json_data["usage"]:
                tokens = json_data["usage"]["total_tokens"]
                total_tokens += tokens
                print(f"本次请求消耗 {tokens} tokens，累计 {total_tokens} tokens")
            
            return content
        except requests.exceptions.RequestException as e:
            print(f"请求异常 ({attempt + 1}/{max_retries}): {str(e)}")
            if attempt < max_retries - 1:
                wait_time = 2 ** attempt  # 指数退避
                print(f"等待 {wait_time} 秒后重试...")
                time.sleep(wait_time)
            else:
                print("达到最大重试次数，放弃请求")
                return None
        except Exception as e:
            print(f"未知错误 ({attempt + 1}/{max_retries}): {str(e)}")
            if attempt < max_retries - 1:
                wait_time = 2 ** attempt
                print(f"等待 {wait_time} 秒后重试...")
                time.sleep(wait_time)
            else:
                print("达到最大重试次数，放弃请求")
                return None
    
    return None

def call_qianwen_model(prompt, text, max_retries=4):
    """调用通义千问模型"""
    headers = {
        "Authorization": f"Bearer {QIANWEN_API_KEY}",
        "Content-Type": "application/json; charset=utf-8"
    }
    system_prompt = """请严格按以下格式生成测试用例：
### 测试用例[编号]：[测试目标]
**优先级**：[高/中/低]
**测试步骤**：
1. [步骤描述]
2. [步骤描述]
**预期结果**：[预期结果描述]"""

    for attempt in range(max_retries):
        try:
            request_data = {
                "model": "qwen-max",
                "input": {
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": f"{prompt}\n相关文本：{text}"}
                    ]
                },
                "parameters": {
                    "temperature": 0.3,
                    "result_format": "message"
                }
            }
            
            # 使用json.dumps确保正确编码
            json_data_str = json.dumps(request_data, ensure_ascii=False)
            
            # 使用data参数而不是json参数，并明确指定编码
            response = requests.post(
                "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation",
                headers=headers,
                data=json_data_str.encode('utf-8'),
                timeout=50
            )
            
            response.raise_for_status()
            json_data = response.json()
            
            # 通义千问API返回格式处理
            content = ""
            if "output" in json_data and "message" in json_data["output"]:
                content = json_data["output"]["message"]["content"]
            
            return {
                "choices": [
                    {
                        "message": {
                            "content": content
                        }
                    }
                ]
            }
        except Exception as e:
            print(f"API请求异常（{str(e)}），重试中 ({attempt + 1}/{max_retries})...")
            time.sleep(5)
    
    print("API请求失败超过最大重试次数")
    return None

def generate_test_cases(requirements, model_name):
    """根据需求生成测试用例"""
    all_test_cases = []
    system_prompt = "你是一位专业的测试工程师，擅长编写详细、全面的测试用例。"
    
    # 对于DeepSeek模型，可能需要调整提示
    if model_name.startswith("deepseek"):
        system_prompt = "你是一位专业的测试工程师，擅长编写详细、全面的测试用例。请严格按照指定格式输出。"
    
    for req in requirements:
        req_id = req["需求ID"]
        req_title = req["标题"]
        req_desc = req["详细描述"]
        priority = req["优先级"]
        parent_req = req.get("父需求", "")
        req_category = req.get("需求分类", DEFAULT_CONFIG["需求分类"])
        iteration = req.get("迭代", DEFAULT_CONFIG["迭代"])
        assignee = req.get("处理人", DEFAULT_CONFIG["处理人"])
        
        print(f"\n处理需求 {req_id}: {req_title[:50]}...")
        
        # 标准化优先级
        std_priority = PRIORITY_MAP.get(priority.lower(), "Middle")
        
        # 构建提示信息
        prompt = f"""
请根据以下需求生成详细的测试用例：

需求ID: {req_id}
需求标题: {req_title}
需求描述: {req_desc}
优先级: {priority}
需求分类: {req_category}
迭代: {iteration}

请生成至少3个测试用例，每个测试用例包括：
1. 测试目标
2. 前置条件
3. 详细的测试步骤
4. 预期结果
5. 优先级（高/中/低）

请严格按照以下格式输出：

### 测试用例1：[测试目标]
**优先级**：[高/中/低]
**前置条件**：[前置条件描述]
**测试步骤**：
1. [步骤1]
2. [步骤2]
...
**预期结果**：[预期结果描述]

### 测试用例2：[测试目标]
...
"""
        
        # 调用AI模型
        if model_name == "gemini-pro":
            # Gemini不支持system角色，将system提示合并到user提示中
            system_prompt = "你是一位专业的测试工程师，擅长编写详细、全面的测试用例。"
            user_prompt = f"{system_prompt}\n\n{prompt}"
            
            response = call_ai_model(
                model_name,
                [
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.7
            )
        else:
            # 其他模型使用标准格式
            response = call_ai_model(
                model_name,
                [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.7
            )
        
        if not response:
            print(f"警告：需求 {req_id} 未能生成测试用例")
            continue
            
        # 解析测试用例
        parsed_cases = parse_test_cases(response, req_id,req_title, parent_req, std_priority)
        all_test_cases.extend(parsed_cases)
        
        print(f"为需求 {req_id} 生成了 {len(parsed_cases)} 条测试用例")
    
    return all_test_cases

def parse_test_cases(response, req_id,req_title, parent_req, std_priority):
    """解析AI生成的测试用例文本"""
    if not response:
        return []

    # 预处理：去除AI对话式的开头
    response = re.sub(r'^.*?(?:### 测试用例|测试用例1|测试用例：)', '', response, flags=re.DOTALL).strip()

    test_cases = []
    case_blocks = re.split(r'###\s+测试用例\d+[:：]', response)
    
    # 移除第一个空块（如果存在）
    if case_blocks and not case_blocks[0].strip():
        case_blocks = case_blocks[1:]
    
    for i, block in enumerate(case_blocks, 1):
        if not block.strip():
            continue
        
        try:
            # 提取测试目标
            title_match = re.match(r'(.+?)(?:\*\*优先级|\n)', block)
            title = title_match.group(1).strip() if title_match else f"测试用例{i}"
            
            # 提取优先级
            priority_match = re.search(r'\*\*优先级\*\*[:：]\s*([高中低]|High|Middle|Low|Nice To Have)', block)
            raw_priority = priority_match.group(1).strip() if priority_match else "中"
            priority = PRIORITY_MAP.get(raw_priority.lower(), std_priority)
            
            # 提取前置条件
            precondition_match = re.search(r'\*\*前置条件\*\*[:：]\s*(.+?)(?=\*\*|\Z)', block, re.DOTALL)
            precondition = precondition_match.group(1).strip() if precondition_match else ""
            
            # 提取测试步骤
            steps_match = re.search(r'\*\*测试步骤\*\*[:：]\s*(.+?)(?=\*\*预期结果|\Z)', block, re.DOTALL)
            steps_text = steps_match.group(1).strip() if steps_match else ""
            
            # 格式化步骤
            steps = []
            for step in re.findall(r'\d+\.\s*(.+?)(?=\d+\.|$)', steps_text, re.DOTALL):
                steps.append(step.strip())
            
            formatted_steps = "\n".join([f"{j+1}. {step}" for j, step in enumerate(steps)])
            
            # 提取预期结果
            expected_match = re.search(r'\*\*预期结果\*\*[:：]\s*(.+?)(?=###|\Z)', block, re.DOTALL)
            expected_result = expected_match.group(1).strip() if expected_match else ""
            
            # 创建测试用例
            case_id = f"TC-{req_id}-{i:02d}"
            test_case = {
                "用例编号": case_id,
                "需求ID": req_id,
                "父需求": parent_req,
                "需求分类": DEFAULT_CONFIG["需求分类"],
                "标题": f"测试-{req_title}-{title}",
                "详细描述": f"**前置条件**：\n{precondition}\n\n**测试步骤**：\n{formatted_steps}\n\n**预期结果**：\n{expected_result}",
                "测试步骤": formatted_steps,
                "预期结果": expected_result,
                "优先级": priority,
                "迭代": DEFAULT_CONFIG["迭代"],
                "处理人": DEFAULT_CONFIG["处理人"]
            }
            
            test_cases.append(test_case)
        except Exception as e:
            print(f"解析测试用例 {i} 时出错: {str(e)}")
    
    return test_cases

def export_to_excel(test_cases, output_file):
    """导出测试用例到Excel"""
    if not test_cases:
        print("错误：没有测试用例可导出")
        return False
    
    try:
        # 确保输出目录存在
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 创建DataFrame
        df = pd.DataFrame(test_cases)
        
        # 创建Excel写入器
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(
                writer,
                index=False,
                sheet_name='测试用例',
                columns=["用例编号", "需求ID", "需求分类", "父需求", "标题", "详细描述", "优先级", "迭代", "处理人"]
            )
            
            # 获取工作表对象
            worksheet = writer.sheets['测试用例']
            
            # 设置列宽
            column_widths = {
                "A": 15,  # 用例编号
                "B": 10,  # 需求ID
                "C": 15,  # 需求分类
                "D": 15,  # 父需求
                "E": 40,  # 标题
                "F": 60,  # 详细描述
                "G": 10,  # 优先级
                "H": 10,  # 迭代
                "I": 15   # 处理人
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置边框样式
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # 设置自动换行和边框
            for row in worksheet.iter_rows(min_row=2, max_col=9, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = border
        
        print(f"成功导出 {len(test_cases)} 条测试用例到 {output_file}")
        return True
    except Exception as e:
        print(f"导出Excel失败: {str(e)}")
        return False

def generate_test_report(test_cases, output_file):
    """生成测试报告"""
    if not test_cases:
        print("错误：没有测试用例可生成报告")
        return False
    
    try:
        # 确保输出目录存在
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 统计信息
        total_cases = len(test_cases)
        priority_stats = {
            "High": len([tc for tc in test_cases if tc["优先级"] == "High"]),
            "Middle": len([tc for tc in test_cases if tc["优先级"] == "Middle"]),
            "Low": len([tc for tc in test_cases if tc["优先级"] == "Low"]),
            "Nice To Have": len([tc for tc in test_cases if tc["优先级"] == "Nice To Have"])
        }
        
        # 按需求ID分组
        req_groups = {}
        for tc in test_cases:
            req_id = tc["需求ID"]
            if req_id not in req_groups:
                req_groups[req_id] = []
            req_groups[req_id].append(tc)
        
        # 创建Excel写入器
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 创建摘要表
            summary_data = {
                "指标": ["总测试用例数", "高优先级", "中优先级", "低优先级", "可选"],
                "数量": [
                    total_cases,
                    priority_stats["High"],
                    priority_stats["Middle"],
                    priority_stats["Low"],
                    priority_stats["Nice To Have"]
                ],
                "百分比": [
                    "100%",
                    f"{priority_stats['High']/total_cases*100:.1f}%" if total_cases > 0 else "0%",
                    f"{priority_stats['Middle']/total_cases*100:.1f}%" if total_cases > 0 else "0%",
                    f"{priority_stats['Low']/total_cases*100:.1f}%" if total_cases > 0 else "0%",
                    f"{priority_stats['Nice To Have']/total_cases*100:.1f}%" if total_cases > 0 else "0%"
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='测试报告摘要', index=False)
            
            # 格式化摘要表
            ws_summary = writer.sheets['测试报告摘要']
            
            # 设置列宽
            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 15
            ws_summary.column_dimensions['C'].width = 15
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for cell in ws_summary[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置单元格边框
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, max_col=3):
                for cell in row:
                    cell.border = border
            
            # 创建需求覆盖表
            coverage_data = []
            for req_id, cases in req_groups.items():
                coverage_data.append({
                    "需求ID": req_id,
                    "测试用例数": len(cases),
                    "高优先级": len([tc for tc in cases if tc["优先级"] == "High"]),
                    "中优先级": len([tc for tc in cases if tc["优先级"] == "Middle"]),
                    "低优先级": len([tc for tc in cases if tc["优先级"] == "Low"]),
                    "可选": len([tc for tc in cases if tc["优先级"] == "Nice To Have"])
                })
            
            coverage_df = pd.DataFrame(coverage_data)
            coverage_df.to_excel(writer, sheet_name='需求覆盖情况', index=False)
            
            # 格式化需求覆盖表
            ws_coverage = writer.sheets['需求覆盖情况']
            
            # 设置列宽
            column_widths = {
                "A": 15,  # 需求ID
                "B": 15,  # 测试用例数
                "C": 15,  # 高优先级
                "D": 15,  # 中优先级
                "E": 15,  # 低优先级
                "F": 15   # 可选
            }
            
            for col, width in column_widths.items():
                ws_coverage.column_dimensions[col].width = width
            
            # 设置表头样式
            for cell in ws_coverage[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置单元格边框
            for row in ws_coverage.iter_rows(min_row=1, max_row=ws_coverage.max_row, max_col=6):
                for cell in row:
                    cell.border = border
            
            # 创建详细测试用例表
            df = pd.DataFrame(test_cases)
            df.to_excel(
                writer,
                sheet_name='详细测试用例',
                index=False,
                columns=["用例编号", "需求ID", "标题", "优先级", "测试步骤", "预期结果"]
            )
            
            # 格式化详细测试用例表
            ws_details = writer.sheets['详细测试用例']
            
            # 设置列宽
            detail_widths = {
                "A": 15,  # 用例编号
                "B": 10,  # 需求ID
                "C": 40,  # 标题
                "D": 10,  # 优先级
                "E": 50,  # 测试步骤
                "F": 40   # 预期结果
            }
            
            for col, width in detail_widths.items():
                ws_details.column_dimensions[col].width = width
            
            # 设置表头样式
            for cell in ws_details[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置单元格样式
            for row in ws_details.iter_rows(min_row=2, max_col=6, max_row=ws_details.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = border
        
        print(f"成功生成测试报告: {output_file}")
        return True
    except Exception as e:
        print(f"生成测试报告失败: {str(e)}")
        return False

def get_available_models():
    """获取可用的模型列表"""
    available_models = []
    
    for model_name, config in MODEL_CONFIGS.items():
        api_key = os.getenv(config["api_key_env"])
        if api_key:
            available_models.append(model_name)
    
    return available_models

def generate_sample_requirements():
    """生成示例需求Excel文件"""
    try:
        # 调用generate_sample_testcase.py中的函数
        from generate_sample_requirements import generate_sample_requirements as generate_sample
        return generate_sample()
    except ImportError as e:
        print(f"无法导入generate_sample_requirements模块: {str(e)}")
        # 如果导入失败，使用内置的示例生成函数
        try:
            # 示例需求数据
            sample_data = [
                {
                    "需求分类": "功能需求",
                    "父需求": "",
                    "标题": "用户登录功能",
                    "详细描述": "系统应支持用户通过用户名和密码进行登录，登录成功后跳转到首页。",
                    "优先级": "高",
                    "迭代": "迭代27",
                    "处理人": ""
                },
                # ... 其他示例数据 ...
            ]
            
            # 创建DataFrame
            df = pd.DataFrame(sample_data)
            
            # 确保目录存在
            os.makedirs("./需求文档", exist_ok=True)
            
            # 保存到Excel
            output_file = "./需求文档/sample_requirements.xlsx"
            df.to_excel(output_file, index=False)
            
            print(f"已生成示例需求文件: {output_file}")
            return output_file
        except Exception as e:
            print(f"生成示例需求文件失败: {str(e)}")
            return None
    except Exception as e:
        print(f"生成示例需求文件失败: {str(e)}")
        return None

def main():
    """主函数"""
    print("=== AITestSuite - 智能测试用例生成器 ===")
    
    # 检查API端点配置
    if not (AI_BASE_URL or AI_API_ENDPOINT):
        print("错误：未配置API端点，请在.env文件中设置AI_BASE_URL或AI_API_ENDPOINT")
        return
    
    # 解析命令行参数
    args = parse_arguments()
    
    # 检查可用模型
    available_models = get_available_models()
    if not available_models:
        print("错误：未找到任何可用的AI模型API密钥，请在.env文件中配置")
        return
    
    print(f"可用模型: {', '.join(available_models)}")
    
    # 获取输入文件
    input_file = args.input
    if not input_file:
        input_file = input("请输入需求Excel文件路径 (默认: 需求文档/sample_requirements.xlsx): ").strip()
        if not input_file:
            input_file = "需求文档/sample_requirements.xlsx"
    
    # 如果文件不存在且是默认文件，尝试生成示例文件
    if not os.path.exists(input_file) and input_file == "需求文档/sample_requirements.xlsx":
        print(f"文件 {input_file} 不存在，将生成示例需求文件")
        input_file = generate_sample_requirements()
        if not input_file:
            return
    elif not os.path.exists(input_file):
        print(f"错误：文件 {input_file} 不存在")
        return
    
    # 选择模型
    model_name = args.model
    if not model_name:
        model_name = input(f"请选择AI模型 (默认: {DEFAULT_MODEL}): ").strip()
        if not model_name:
            model_name = DEFAULT_MODEL
    
    if model_name not in MODEL_CONFIGS:
        print(f"错误：不支持的模型 '{model_name}'，将使用默认模型 '{DEFAULT_MODEL}'")
        model_name = DEFAULT_MODEL
    
    if model_name not in available_models:
        print(f"错误：模型 '{model_name}' 的API密钥未配置")
        return
    
    # 设置输出目录
    test_cases_dir = args.output_dir
    test_report_dir = args.report_dir
    os.makedirs(test_cases_dir, exist_ok=True)
    os.makedirs(test_report_dir, exist_ok=True)
    
    # 生成时间戳
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    # 设置输出文件
    test_cases_file = f"{test_cases_dir}/TestCases_{model_name}_{timestamp}.xlsx"
    test_report_file = f"{test_report_dir}/TestReport_{model_name}_{timestamp}.xlsx"
    
    # 读取需求
    print(f"\n读取需求文件: {input_file}")
    requirements = read_excel_requirements(input_file)
    
    if not requirements:
        print("错误：无法读取需求数据")
        return
    
    print(f"成功读取 {len(requirements)} 条需求")
    
    # 生成测试用例
    print(f"\n使用模型 {model_name} 生成测试用例...")
    all_test_cases = generate_test_cases(requirements, model_name)
    
    # 检查是否生成了测试用例
    if not all_test_cases:
        print("错误：未生成任何测试用例")
        return
    
    # 导出测试用例
    print(f"\n导出测试用例到: {test_cases_file}")
    if export_to_excel(all_test_cases, test_cases_file):
        # 生成测试报告
        print(f"\n生成测试报告: {test_report_file}")
        generate_test_report(all_test_cases, test_report_file)
    
    print("\n处理完成！")
    print(f"- 测试用例文件: {test_cases_file}")
    print(f"- 测试报告文件: {test_report_file}")

if __name__ == "__main__":
    main()
