import os
import json
import time
import requests
import re
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from dotenv import load_dotenv

class AITestSuiteUtils:
    def __init__(self):
        # 加载环境变量
        load_dotenv()
        
        # 初始化配置
        self.AI_BASE_URL = os.getenv("AI_BASE_URL")
        self.AI_API_ENDPOINT = os.getenv("AI_API_ENDPOINT")
        self.AI_API_KEY = os.getenv("AI_API_KEY")
        self.MODEL_NAME = os.getenv("MODEL_NAME", "default")
        self.API_ENDPOINT = self.AI_BASE_URL or self.AI_API_ENDPOINT
        
        # 初始化模型配置
        self.MODEL_CONFIGS = {
            "default": {
                "api_key_env": "AI_API_KEY",
                "endpoint": self.API_ENDPOINT,
                "headers": lambda key: {
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json; charset=utf-8"
                },
                "payload": lambda messages, temperature: {
                    "model": self.MODEL_NAME,
                    "messages": messages,
                    "temperature": temperature
                },
                "response_parser": lambda json_data: json_data["choices"][0]["message"]["content"] if "choices" in json_data and len(json_data["choices"]) > 0 else ""
            }
        }
        
        # 添加其他模型配置
        self._add_model_configs()
        
        # 默认配置
        self.DEFAULT_MODEL = "default"
        self.DEFAULT_CONFIG = {
            "需求分类": "测试需求",
            "迭代": "迭代",
            "处理人": ""
        }
        self.PRIORITY_MAP = {
            "高": "High", "中": "Middle", "低": "Low", "可选": "Nice To Have",
            "high": "High", "middle": "Middle", "low": "Low", "nice to have": "Nice To Have"
        }

    def _add_model_configs(self):
        """添加其他模型配置"""
        if os.getenv("QIANWEN_API_KEY"):
            self.MODEL_CONFIGS["qianwen"] = {
                "api_key_env": "QIANWEN_API_KEY",
                "endpoint": "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation",
                "headers": lambda key: {
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json; charset=utf-8"
                },
                "payload": lambda messages, temperature: {
                    "model": "qwen-max",
                    "input": {"messages": messages},
                    "parameters": {"temperature": temperature, "result_format": "message"}
                },
                "response_parser": lambda json_data: json_data["output"]["message"]["content"] if "output" in json_data and "message" in json_data["output"] else ""
            }
        
        if os.getenv("GEMINI_API_KEY"):
            self.MODEL_CONFIGS["mygemini"] = {
                "api_key_env": "GEMINI_API_KEY",
                "endpoint": os.getenv("GEMINI_BASE_URL"),
                "headers": lambda key: {
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json"
                },
                "payload": lambda messages, temperature: {
                    "model": os.getenv("GEMINI_MODEL_NAME"),
                    "messages": messages,
                    "temperature": temperature
                },
                "response_parser": lambda json_data: json_data["choices"][0]["message"]["content"] if "choices" in json_data and len(json_data["choices"]) > 0 else ""
            }

    def generate_sample_requirements(self):
        """生成示例需求Excel文件"""
        try:
            # 调用generate_sample_testcase.py中的函数
            from generate_sample_requirements import generate_sample_requirements as generate_sample
            return generate_sample()
        except ImportError as e:
            print(f"无法导入generate_sample_requirements模块: {str(e)}")
            # 如果导入失败，使用内置的示例生成函数
            try:
                # 示例需求数据
                sample_data = [
                    {
                        "需求分类": "功能需求",
                        "父需求": "",
                        "标题": "用户登录功能",
                        "详细描述": "系统应支持用户通过用户名和密码进行登录，登录成功后跳转到首页。",
                        "优先级": "高",
                        "迭代": "迭代",
                        "处理人": ""
                    },
                    # ... 其他示例数据 ...
                ]
            
                # 创建DataFrame
                df = pd.DataFrame(sample_data)
            
                # 确保目录存在
                os.makedirs("./需求文档", exist_ok=True)
            
                # 保存到Excel
                output_file = "./需求文档/sample_requirements.xlsx"
                df.to_excel(output_file, index=False)
            
                print(f"已生成示例需求文件: {output_file}")
                return output_file
            except Exception as e:
                print(f"生成示例需求文件失败: {str(e)}")
                return None
        except Exception as e:
            print(f"生成示例需求文件失败: {str(e)}")
            return None

    def generate_pdf_requirements(self, output_file="./需求文档/API文档示例.pdf"):

        """生成示例PDF需求文档"""
        try:
            # 导入pdf_create_sample_requirements模块
            from pdf_create_sample_requirements import create_pdf_document
            # 调用生成PDF的函数
            create_pdf_document(output_file)
            print(f"示例PDF需求文档已生成: {output_file}")
            return True
        except ImportError as e:
            print(f"无法导入pdf_create_sample_requirements模块: {str(e)}")
            # 如果导入失败，使用内置的PDF生成逻辑
            try:
                # 这里可以添加内置的PDF生成逻辑
                print("使用内置逻辑生成PDF文档")
                return False
            except Exception as e:
                print(f"生成PDF需求文档失败: {str(e)}")
                return False

    def read_excel_requirements(self, file_path):
        """读取Excel需求文档"""
        try:
            df = pd.read_excel(file_path)
            
            # 检查必要的列是否存在
            required_columns = ["标题", "详细描述"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                print(f"错误：Excel文件缺少必要的列: {', '.join(missing_columns)}")
                return None
            
            # 如果没有需求ID列，添加自动生成的ID
            if "需求ID" not in df.columns:
                df["需求ID"] = [f"REQ{i+1:03d}" for i in range(len(df))]
            
            # 如果没有优先级列，添加默认值
            if "优先级" not in df.columns:
                df["优先级"] = "中"
            
            # 如果没有父需求列，添加空值
            if "父需求" not in df.columns:
                df["父需求"] = ""
                
            # 如果没有需求分类列，添加默认值
            if "需求分类" not in df.columns:
                df["需求分类"] = self.DEFAULT_CONFIG["需求分类"]
                
            # 如果没有迭代列，添加默认值
            if "迭代" not in df.columns:
                df["迭代"] = self.DEFAULT_CONFIG["迭代"]
                
            # 如果没有处理人列，添加默认值
            if "处理人" not in df.columns:
                df["处理人"] = self.DEFAULT_CONFIG["处理人"]
            
            # 转换为字典列表
            requirements = df.to_dict('records')
            return requirements
        except Exception as e:
            print(f"读取Excel文件失败: {str(e)}")
            return None

    def call_ai_model(self, model_name, messages, max_retries=3, temperature=0.3):
        """调用AI模型生成内容"""
        if model_name not in self.MODEL_CONFIGS:
            print(f"错误：不支持的模型 '{model_name}'，将使用默认模型 '{self.DEFAULT_MODEL}'")
            model_name = self.DEFAULT_MODEL
        
        model_config = self.MODEL_CONFIGS[model_name]
        api_key = os.getenv(model_config["api_key_env"])
        
        if not api_key:
            print(f"错误：未设置 {model_config['api_key_env']} 环境变量")
            return None
        
        headers = model_config["headers"](api_key)
        endpoint = model_config["endpoint"]
        
        if not endpoint:
            print(f"错误：未设置API端点，请在.env文件中配置AI_BASE_URL或AI_API_ENDPOINT")
            return None
        
        total_tokens = 0
        
        for attempt in range(max_retries):
            try:
                payload = model_config["payload"](messages, temperature)
                
                # 将payload转换为JSON字符串，确保正确处理中文
                json_data_str = json.dumps(payload, ensure_ascii=False)
                
                # 准备请求参数
                request_kwargs = {
                    "headers": headers,
                    "data": json_data_str.encode('utf-8'),
                    "timeout": 60
                }
                
                # 如果有URL参数，添加到请求中
                if "url_params" in model_config:
                    url_params = model_config["url_params"](api_key)
                    endpoint = f"{endpoint}?{'&'.join([f'{k}={v}' for k, v in url_params.items()])}"
                
                # 发送请求
                print(f"正在调用API: {endpoint}")
                response = requests.post(endpoint, **request_kwargs)
                
                # 检查响应状态
                response.raise_for_status()
                
                # 解析响应
                json_data = response.json()
                
                # 使用模型特定的解析器提取内容
                content = model_config["response_parser"](json_data)
                
                if not content:
                    print(f"警告：模型 {model_name} 返回的内容为空")
                    
                if "usage" in json_data and "total_tokens" in json_data["usage"]:
                    tokens = json_data["usage"]["total_tokens"]
                    total_tokens += tokens
                    print(f"本次请求消耗 {tokens} tokens，累计 {total_tokens} tokens")
                
                return content
            except requests.exceptions.RequestException as e:
                print(f"请求异常 ({attempt + 1}/{max_retries}): {str(e)}")
                if attempt < max_retries - 1:
                    wait_time = 2 ** attempt  # 指数退避
                    print(f"等待 {wait_time} 秒后重试...")
                    time.sleep(wait_time)
                else:
                    print("达到最大重试次数，放弃请求")
                    return None
            except Exception as e:
                print(f"未知错误 ({attempt + 1}/{max_retries}): {str(e)}")
                if attempt < max_retries - 1:
                    wait_time = 2 ** attempt
                    print(f"等待 {wait_time} 秒后重试...")
                    time.sleep(wait_time)
                else:
                    print("达到最大重试次数，放弃请求")
                    return None
        
        return None

    def generate_test_cases(self, requirements, model_name):
        """根据需求生成测试用例"""
        all_test_cases = []
        system_prompt = "你是一位专业的测试工程师，擅长编写详细、全面的测试用例。"
        
        # 对于DeepSeek模型，可能需要调整提示
        if model_name.startswith("deepseek"):
            system_prompt = "你是一位专业的测试工程师，擅长编写详细、全面的测试用例。请严格按照指定格式输出。"
        
        for req in requirements:
            req_id = req["需求ID"]
            req_title = req["标题"]
            req_desc = req["详细描述"]
            priority = req["优先级"]
            parent_req = req.get("父需求", "")
            req_category = req.get("需求分类", self.DEFAULT_CONFIG["需求分类"])
            iteration = req.get("迭代", self.DEFAULT_CONFIG["迭代"])
            assignee = req.get("处理人", self.DEFAULT_CONFIG["处理人"])
            
            print(f"\n处理需求 {req_id}: {req_title[:50]}...")
            
            # 标准化优先级
            std_priority = self.PRIORITY_MAP.get(priority.lower(), "Middle")
            
            # 构建提示信息
            prompt = f"""
请根据以下需求生成详细的测试用例：

需求ID: {req_id}
需求标题: {req_title}
需求描述: {req_desc}
优先级: {priority}
需求分类: {req_category}
迭代: {iteration}

请生成至少3个测试用例，每个测试用例包括：
1. 测试目标
2. 前置条件
3. 详细的测试步骤
4. 预期结果
5. 优先级（高/中/低）

请严格按照以下格式输出：

### 测试用例1：[测试目标]
**优先级**：[高/中/低]
**前置条件**：[前置条件描述]
**测试步骤**：
1. [步骤1]
2. [步骤2]
...
**预期结果**：[预期结果描述]

### 测试用例2：[测试目标]
...
"""
            
            # 调用AI模型
            if model_name == "gemini-pro":
                # Gemini不支持system角色，将system提示合并到user提示中
                system_prompt = "你是一位专业的测试工程师，擅长编写详细、全面的测试用例。"
                user_prompt = f"{system_prompt}\n\n{prompt}"
                
                response = self.call_ai_model(
                    model_name,
                    [
                        {"role": "user", "content": user_prompt}
                    ],
                    temperature=0.7
                )
            else:
                # 其他模型使用标准格式
                response = self.call_ai_model(
                    model_name,
                    [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.7
                )
            
            if not response:
                print(f"警告：需求 {req_id} 未能生成测试用例")
                continue
                
            # 解析测试用例
            parsed_cases = self.parse_test_cases(response, req_id, req_title, parent_req, std_priority)
            all_test_cases.extend(parsed_cases)
            
            print(f"为需求 {req_id} 生成了 {len(parsed_cases)} 条测试用例")
        
        return all_test_cases

    def parse_test_cases(self, response, req_id, req_title, parent_req, std_priority):
        """解析AI生成的测试用例文本"""
        if not response:
            return []

        # 预处理：去除AI对话式的开头
        response = re.sub(r'^.*?(?:### 测试用例|测试用例1|测试用例：)', '', response, flags=re.DOTALL).strip()

        test_cases = []
        case_blocks = re.split(r'###\s+测试用例\d+[:：]', response)
        
        # 移除第一个空块（如果存在）
        if case_blocks and not case_blocks[0].strip():
            case_blocks = case_blocks[1:]
        
        for i, block in enumerate(case_blocks, 1):
            if not block.strip():
                continue
            
            try:
                # 提取测试目标
                title_match = re.match(r'(.+?)(?:\*\*优先级|\n)', block)
                title = title_match.group(1).strip() if title_match else f"测试用例{i}"
                
                # 提取优先级
                priority_match = re.search(r'\*\*优先级\*\*[:：]\s*([高中低]|High|Middle|Low|Nice To Have)', block)
                raw_priority = priority_match.group(1).strip() if priority_match else "中"
                priority = self.PRIORITY_MAP.get(raw_priority.lower(), std_priority)
                
                # 提取前置条件
                precondition_match = re.search(r'\*\*前置条件\*\*[:：]\s*(.+?)(?=\*\*|\Z)', block, re.DOTALL)
                precondition = precondition_match.group(1).strip() if precondition_match else ""
                
                # 提取测试步骤
                steps_match = re.search(r'\*\*测试步骤\*\*[:：]\s*(.+?)(?=\*\*预期结果|\Z)', block, re.DOTALL)
                steps_text = steps_match.group(1).strip() if steps_match else ""
                
                # 格式化步骤
                steps = []
                for step in re.findall(r'\d+\.\s*(.+?)(?=\d+\.|$)', steps_text, re.DOTALL):
                    steps.append(step.strip())
                
                formatted_steps = "\n".join([f"{j+1}. {step}" for j, step in enumerate(steps)])
                
                # 提取预期结果
                expected_match = re.search(r'\*\*预期结果\*\*[:：]\s*(.+?)(?=###|\Z)', block, re.DOTALL)
                expected_result = expected_match.group(1).strip() if expected_match else ""
                
                # 创建测试用例
                case_id = f"TC-{req_id}-{i:02d}"
                test_case = {
                    "用例编号": case_id,
                    "需求ID": req_id,
                    "父需求": parent_req,
                    "需求分类": self.DEFAULT_CONFIG["需求分类"],
                    "标题": f"测试-{req_title}-{title}",
                    "详细描述": f"**前置条件**：\n{precondition}\n\n**测试步骤**：\n{formatted_steps}\n\n**预期结果**：\n{expected_result}",
                    "测试步骤": formatted_steps,
                    "预期结果": expected_result,
                    "优先级": priority,
                    "迭代": self.DEFAULT_CONFIG["迭代"],
                    "处理人": self.DEFAULT_CONFIG["处理人"]
                }
                test_cases.append(test_case)
            except Exception as e:
                print(f"解析测试用例失败: {str(e)}")
                continue
        
        return test_cases

    def export_to_excel(self, test_cases, output_file):
        """导出测试用例到Excel"""
        if not test_cases:
            print("错误：没有测试用例可导出")
            return False
        
        try:
            # 确保输出目录存在
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 将测试用例转换为DataFrame
            df = pd.DataFrame(test_cases)
            
            # 确保必要的列存在
            required_columns = ["需求ID", "标题", "测试步骤", "预期结果", "优先级"]
            for col in required_columns:
                if col not in df.columns:
                    df[col] = ""  # 添加缺失列并填充空值
            
            # 导出到Excel
            df.to_excel(output_file, index=False, engine="openpyxl")
            print(f"测试用例已成功导出到: {output_file}")
            return True
        except Exception as e:
            print(f"导出测试用例失败: {str(e)}")
            return False

    def generate_test_report(self, test_cases, output_file):
        """生成测试报告"""
        if not test_cases:
            print("错误：没有测试用例可生成报告")
            return False
    
        try:
            # 确保输出目录存在
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
        
            # 统计信息
            total_cases = len(test_cases)
            priority_stats = {
                "High": len([tc for tc in test_cases if tc["优先级"] == "High"]),
                "Middle": len([tc for tc in test_cases if tc["优先级"] == "Middle"]),
                "Low": len([tc for tc in test_cases if tc["优先级"] == "Low"]),
                "Nice To Have": len([tc for tc in test_cases if tc["优先级"] == "Nice To Have"])
            }
        
            # 按需求ID分组
            req_groups = {}
            for tc in test_cases:
                req_id = tc["需求ID"]
                if req_id not in req_groups:
                    req_groups[req_id] = []
                req_groups[req_id].append(tc)
        
            # 创建Excel写入器
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 创建摘要表
                summary_data = {
                    "指标": ["总测试用例数", "高优先级", "中优先级", "低优先级", "可选"],
                    "数量": [
                        total_cases,
                        priority_stats["High"],
                        priority_stats["Middle"],
                        priority_stats["Low"],
                        priority_stats["Nice To Have"]
                    ],
                    "百分比": [
                        "100%",
                        f"{priority_stats['High']/total_cases*100:.1f}%" if total_cases > 0 else "0%",
                        f"{priority_stats['Middle']/total_cases*100:.1f}%" if total_cases > 0 else "0%",
                        f"{priority_stats['Low']/total_cases*100:.1f}%" if total_cases > 0 else "0%",
                        f"{priority_stats['Nice To Have']/total_cases*100:.1f}%" if total_cases > 0 else "0%"
                    ]
                }
            
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='测试报告摘要', index=False)
            
                # 格式化摘要表
                ws_summary = writer.sheets['测试报告摘要']
            
                # 设置列宽
                ws_summary.column_dimensions['A'].width = 20
                ws_summary.column_dimensions['B'].width = 15
                ws_summary.column_dimensions['C'].width = 15
            
                # 设置表头样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
                for cell in ws_summary[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
                # 设置单元格边框
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
                for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, max_col=3):
                    for cell in row:
                        cell.border = border
            
                # 创建需求覆盖表
                coverage_data = []
                for req_id, cases in req_groups.items():
                    coverage_data.append({
                        "需求ID": req_id,
                        "测试用例数": len(cases),
                        "高优先级": len([tc for tc in cases if tc["优先级"] == "High"]),
                        "中优先级": len([tc for tc in cases if tc["优先级"] == "Middle"]),
                        "低优先级": len([tc for tc in cases if tc["优先级"] == "Low"]),
                        "可选": len([tc for tc in cases if tc["优先级"] == "Nice To Have"])
                    })
            
                coverage_df = pd.DataFrame(coverage_data)
                coverage_df.to_excel(writer, sheet_name='需求覆盖情况', index=False)
            
                # 格式化需求覆盖表
                ws_coverage = writer.sheets['需求覆盖情况']
            
                # 设置列宽
                column_widths = {
                    "A": 15,  # 需求ID
                    "B": 15,  # 测试用例数
                    "C": 15,  # 高优先级
                    "D": 15,  # 中优先级
                    "E": 15,  # 低优先级
                    "F": 15   # 可选
                }
            
                for col, width in column_widths.items():
                    ws_coverage.column_dimensions[col].width = width
            
                # 设置表头样式
                for cell in ws_coverage[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
                # 设置单元格边框
                for row in ws_coverage.iter_rows(min_row=1, max_row=ws_coverage.max_row, max_col=6):
                    for cell in row:
                        cell.border = border
            
                # 创建详细测试用例表
                df = pd.DataFrame(test_cases)
                df.to_excel(
                    writer,
                    sheet_name='详细测试用例',
                    index=False,
                    columns=["用例编号", "需求ID", "标题", "优先级", "测试步骤", "预期结果"]
                )
            
                # 格式化详细测试用例表
                ws_details = writer.sheets['详细测试用例']
            
                # 设置列宽
                detail_widths = {
                    "A": 15,  # 用例编号
                    "B": 10,  # 需求ID
                    "C": 40,  # 标题
                    "D": 10,  # 优先级
                    "E": 50,  # 测试步骤
                    "F": 40   # 预期结果
                }
            
                for col, width in detail_widths.items():
                    ws_details.column_dimensions[col].width = width
            
                # 设置表头样式
                for cell in ws_details[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
                # 设置单元格样式
                for row in ws_details.iter_rows(min_row=2, max_col=6, max_row=ws_details.max_row):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = border
        
            print(f"成功生成测试报告: {output_file}")
            return True
        except Exception as e:
            print(f"生成测试报告失败: {str(e)}")
            return False
